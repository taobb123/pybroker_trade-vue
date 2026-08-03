#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
对早间涨停股票做主题分类（方案A：仅基于 industry + name）。

固定优先级：
AI算力 > 新能源 > 化工 > 有色金属 > 其它
"""

from __future__ import annotations

import argparse
import os
from typing import Iterable

import pandas as pd
from openpyxl.utils import get_column_letter


DISPLAY_PRIORITY = ("AI算力", "新能源", "化工", "有色金属", "高端制造", "其它")
SORT_PRIORITY = DISPLAY_PRIORITY

AI_INDUSTRY_KEYWORDS = (
    "通信设备",
    "it服务",
    "计算机",
    "软件",
    "互联网",
    "元件",
    "半导体",
    "光学光电",
    "电子化学",
)

AI_NAME_KEYWORDS = (
    "ai",
    "算力",
    "数据中心",
    "服务器",
    "光模块",
    "cpo",
    "液冷",
    "云计算",
    "大数据",
    "人工智能",
    "互联",
    "光电",
)

NEW_ENERGY_INDUSTRY_KEYWORDS = (
    "光伏",
    "风电",
    "锂电",
    "电池",
    "储能",
    "新能源",
    "电力设备",
)

NEW_ENERGY_NAME_KEYWORDS = (
    "光伏",
    "风电",
    "锂电",
    "钠电",
    "储能",
    "电池",
    "充电桩",
    "逆变器",
    "组件",
)

CHEMICAL_INDUSTRY_KEYWORDS = (
    "化工",
    "化学原料",
    "化学制品",
    "农化",
    "化肥",
    "农药",
    "化纤",
    "橡胶",
    "塑料",
)

CHEMICAL_NAME_KEYWORDS = (
    "化工",
    "新材",
    "树脂",
    "纯碱",
    "烧碱",
    "磷化",
    "硅料",
    "tma",
    "pta",
    "pe",
    "pp",
)

NONFERROUS_INDUSTRY_KEYWORDS = (
    "有色",
    "小金属",
    "工业金属",
    "金属新材",
    "稀有金属",
)

NONFERROUS_NAME_KEYWORDS = (
    "稀土",
    "钨",
    "钼",
    "锗",
    "铜",
    "铝",
    "锌",
    "镍",
    "钴",
    "锂",
    "铅",
    "锡",
)

ADV_MANUFACTURING_INDUSTRY_KEYWORDS = (
    "高端制造",
    "专用设备",
    "通用设备",
    "自动化",
    "航空装备",
    "航海装备",
    "军工",
    "地面兵装",
    "汽车零部件",
    "电网设备",
    "光伏设备",
    "工程机械",
)

ADV_MANUFACTURING_NAME_KEYWORDS = (
    "装备",
    "制造",
    "机床",
    "机器人",
    "电缆",
    "重机",
    "数控",
)


def _contains_any(text: str, keywords: Iterable[str]) -> bool:
    t = (text or "").lower()
    return any(k.lower() in t for k in keywords)


def classify_row(industry: str, name: str) -> str:
    ind = (industry or "").strip().lower()
    nm = (name or "").strip().lower()
    if _contains_any(ind, AI_INDUSTRY_KEYWORDS) or _contains_any(nm, AI_NAME_KEYWORDS):
        return "AI算力"

    if _contains_any(ind, NEW_ENERGY_INDUSTRY_KEYWORDS) or _contains_any(nm, NEW_ENERGY_NAME_KEYWORDS):
        return "新能源"

    if _contains_any(ind, CHEMICAL_INDUSTRY_KEYWORDS) or _contains_any(nm, CHEMICAL_NAME_KEYWORDS):
        return "化工"

    if _contains_any(ind, NONFERROUS_INDUSTRY_KEYWORDS) or _contains_any(nm, NONFERROUS_NAME_KEYWORDS):
        return "有色金属"

    if _contains_any(ind, ADV_MANUFACTURING_INDUSTRY_KEYWORDS) or _contains_any(nm, ADV_MANUFACTURING_NAME_KEYWORDS):
        return "高端制造"

    return "其它"


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(c).strip() for c in out.columns]
    return out


def _pick_col(df: pd.DataFrame, candidates: tuple[str, ...]) -> str:
    col_map = {str(c).strip().lower(): str(c) for c in df.columns}
    for c in candidates:
        k = c.strip().lower()
        if k in col_map:
            return col_map[k]
    return ""


def _read_table(input_path: str) -> pd.DataFrame:
    ext = os.path.splitext(input_path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return pd.read_excel(input_path)
    return pd.read_csv(input_path, dtype={"ts_code": "string", "code6": "string", "symbol": "string"})


def _load_input_table(input_path: str) -> pd.DataFrame:
    df = _normalize_columns(_read_table(input_path))

    code_col = _pick_col(df, ("code6", "代码", "code", "symbol", "ts_code"))
    name_col = _pick_col(df, ("name", "stock_name", "名称", "股票名称"))
    industry_col = _pick_col(df, ("industry", "所属行业", "行业"))

    if code_col:
        normalized_code = df[code_col].astype(str).str.extract(r"(\d+)", expand=False).fillna("").str.zfill(6).str[-6:]
        df["code6"] = normalized_code
        # 对常见纯数字代码列同步补齐，避免输出里出现去掉前导 0 的情况。
        if code_col.strip().lower() in {"代码", "code", "symbol", "code6"}:
            df[code_col] = normalized_code
    if name_col:
        df["name"] = df[name_col].astype(str).str.strip()
    if industry_col:
        df["industry"] = df[industry_col].astype(str).str.strip()

    has_fields = ("industry" in df.columns) or ("name" in df.columns)
    if has_fields:
        return df

    # 兼容纯代码列表文本（每行一个股票代码）。
    codes = []
    with open(input_path, "r", encoding="utf-8-sig") as f:
        for line in f:
            raw = line.strip()
            if not raw:
                continue
            code = "".join(ch for ch in raw if ch.isdigit()).zfill(6)[-6:]
            if code and code not in codes:
                codes.append(code)
    return pd.DataFrame({"code6": codes, "industry": "", "name": "", "stock_name": ""})


def run(input_path: str, output_path: str) -> None:
    df = _load_input_table(input_path)
    if "industry" not in df.columns:
        df["industry"] = ""
    if "name" not in df.columns:
        df["name"] = ""
    if "code6" in df.columns:
        df["code6"] = df["code6"].astype(str).str.extract(r"(\d+)", expand=False).fillna("").str.zfill(6)

    df["theme"] = df.apply(lambda r: classify_row(str(r.get("industry", "")), str(r.get("name", ""))), axis=1)
    df["theme"] = pd.Categorical(df["theme"], categories=list(SORT_PRIORITY), ordered=True)
    sort_cols = ["theme"] + (["first_time_hms"] if "first_time_hms" in df.columns else [])
    asc = [True] * len(sort_cols)
    df = df.sort_values(by=sort_cols, ascending=asc, na_position="last").reset_index(drop=True)
    df["theme"] = df["theme"].astype(str)
    out_ext = os.path.splitext(output_path)[1].lower()
    if out_ext in (".xlsx", ".xls"):
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            ws = writer.sheets["Sheet1"]
            code_cols = [c for c in ("代码", "code", "symbol", "code6") if c in df.columns]
            col_index_map = {str(c): i + 1 for i, c in enumerate(df.columns)}
            for col_name in code_cols:
                col_idx = col_index_map[col_name]
                col_letter = get_column_letter(col_idx)
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws[f"{col_letter}{row_idx}"]
                    if cell.value is None:
                        continue
                    digits = "".join(ch for ch in str(cell.value) if ch.isdigit())
                    if not digits:
                        continue
                    cell.value = digits.zfill(6)[-6:]
                    cell.number_format = "@"
    else:
        df.to_csv(output_path, index=False, encoding="utf-8-sig")

    counts = df["theme"].value_counts().to_dict()
    print(f"输入: {os.path.abspath(input_path)}")
    print(f"输出: {os.path.abspath(output_path)}")
    print("分类统计:")
    for key in DISPLAY_PRIORITY:
        print(f"  - {key}: {int(counts.get(key, 0))}")


def main() -> None:
    desktop_dir = r"C:\Users\111\Desktop"
    parser = argparse.ArgumentParser(description="分类早间涨停股票主题（AI算力/新能源/化工/有色金属/其它）")
    parser.add_argument(
        "--input",
        default=os.path.join(desktop_dir, "in_all.xlsx"),
        help="输入路径（支持 xlsx/csv/txt）",
    )
    parser.add_argument(
        "--output",
        default=os.path.join(desktop_dir, "out.txt.xlsx"),
        help="输出路径（默认 out.txt.xlsx）",
    )
    args = parser.parse_args()

    input_path = args.input
    output_path = args.output
    run(input_path=input_path, output_path=output_path)


if __name__ == "__main__":
    main()
